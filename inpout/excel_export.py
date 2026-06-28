from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from constants import path_to_export
from models import Kanji
from services.info import (
    ShowKanjiWithPronouncesAndMeaning,
    SafeJoin
)

def put_kanji_in_excel(kanjiList):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "KanjiList"
    sheet = fill_head_columns(sheet)
    row = 2
    for kanji in kanjiList:
        kanjiObject = Kanji()
        kanjiObject.hieroglyph = kanji[0]
        kanjiDict = ShowKanjiWithPronouncesAndMeaning(kanjiObject)
        sheet = fill_columns(sheet, row, kanjiDict)
        row += 1
    FileFormat(sheet)
    wb.save(path_to_export)
    return path_to_export

def make_all_of_radical_table(radicalKanjiDict):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "RadicalKanjiList"
    sheet.cell(row = 1, column = 1).value = "Radical"
    sheet.cell(row = 1, column = 2).value = "Kanjies"
    row = 2
    for radical, kanji in radicalKanjiDict.items():
        sheet.cell(row = row, column = 1).value = radical
        sheet.cell(row = row, column = 2).value = kanji
        row += 1
    FileFormat(sheet)
    wb.save(path_to_export)
    return path_to_export

def fill_head_columns(ws):
    ws.cell(row = 1, column = 1).value = "Kanji"
    ws.cell(row = 1, column = 2).value = "Onyomi"
    ws.cell(row = 1, column = 3).value = "Ony_latin"
    ws.cell(row = 1, column = 4).value = "Kunyomi"
    ws.cell(row = 1, column = 5).value = "Kuny_latin"
    ws.cell(row = 1, column = 6).value = "Eng"
    ws.cell(row = 1, column = 7).value = "Rus"
    return ws

def fill_columns(ws, row, kanjiDict):
    ws.cell(row = row, column = 1).value = kanjiDict["kanji"]
    ws.cell(row = row, column = 2).value = SafeJoin(kanjiDict["onyomi"])
    ws.cell(row = row, column = 3).value = SafeJoin(kanjiDict["ony_latin"])
    ws.cell(row = row, column = 4).value = SafeJoin(kanjiDict["kunyomi"])
    ws.cell(row = row, column = 5).value = SafeJoin(kanjiDict["kuny_latin"])
    ws.cell(row = row, column = 6).value = SafeJoin(kanjiDict["eng"])
    ws.cell(row = row, column = 7).value = SafeJoin(kanjiDict["rus"])
    return ws

def FileFormat(ws):
    header_font = Font(bold = True)
    header_fill = PatternFill(
        fill_type = "solid",
        fgColor = "D9EAD3"
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal = "center",
                vertical = "center",
                wrap_text = True
            )