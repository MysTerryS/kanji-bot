from openpyxl import load_workbook
from os import remove
from constants import path_to_export
from services.import_service import (
    SendToBase,
    SendRadicalToBase
)
from services.parser import (
    SaveValue
)
from services.search import (
    MakeLinkWithRadical
)

def load_excel(file):
    wb = load_workbook(file)
    ws = wb.active
    return wb, ws

def read_excel(file, isRadical):
    wb, ws = load_excel(file)
    row = 2
    try:
        while ws.cell(row = row, column = 1).value != None:
            kanjiDict = read_kanji(ws, row)
            SendToBase(kanjiDict)
            if isRadical:
                SendRadicalToBase(kanjiDict["kanji"])
            row += 1
        wb.close()
        remove(file)
        return "Successful"
    except Exception as ex:
        return ex
    
def read_kanji(ws, row):
    kanjiDict = dict()
    column = 1
    while column <= 8:
        key = ws.cell(row = 1, column = column).value.lower()
        value = ws.cell(row = row, column = column).value
        kanjiDict[key] = SaveValue(key, value)
        column += 1
    return kanjiDict

def read_links_excel(file):
    wb, ws = load_excel(file)
    row = 2
    try:
        while ws.cell(row = row, column = 1).value != None:
            text = "link " + ws.cell(row = row, column = 1).value + " " + ws.cell(row = row, column = 2).value
            MakeLinkWithRadical(text)
            row += 1
        wb.close()
        return "Successful"
    except Exception as ex:
        return ex